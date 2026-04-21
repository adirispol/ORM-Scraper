import streamlit as st
import pandas as pd
import requests
import time
import io
import html
import os
import re
from datetime import datetime, timedelta, timezone, date
import urllib.parse
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PAGE CONFIG — must be FIRST streamlit call
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
st.set_page_config(
    page_title="Polaris Brand Intelligence",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CONSTANTS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
APIFY_BASE = "https://api.apify.com/v2"
IST = timezone(timedelta(hours=5, minutes=30))

ACTORS = {
    "linkedin":  "supreme_coder~linkedin-post",
    "twitter":   "xquik~x-tweet-scraper",
    "instagram": "apify~instagram-hashtag-scraper",
    "facebook":  "apify~facebook-posts-scraper",
    "quora":     "tri_angle~quora-scraper",
    "reddit":    "trudax~reddit-scraper",
    "youtube":   "streamers~youtube-scraper",
}

COSTS = {
    "linkedin":  0.005,
    "twitter":   0.00015,
    "instagram": 0.004,
    "facebook":  0.004,
    "quora":     0.005,
    "reddit":    0.005,
    "youtube":   0.01,
}

PLATFORM_COLORS = {
    "LinkedIn":  "#0A66C2",
    "Twitter":   "#1DA1F2",
    "Instagram": "#E1306C",
    "Facebook":  "#1877F2",
    "Quora":     "#B92B27",
    "Reddit":    "#FF4500",
    "YouTube":   "#FF0000",
    "Web/News":  "#099B5A",
}

COMPETITORS = ["scaler", "newton school", "upgrad", "great learning", "iit bombay", "iit delhi", "iit madras"]
BRAND_KEYWORDS = ["polaris", "polariscampus", "polaris school of technology", "pst bengaluru"]

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# STYLES
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;500;600;700;800&family=Playfair+Display:wght@700&display=swap');
:root {
    --pg: #FFCC3F; --po: #FF9A3F; --pa: #D88100;
    --pb: #020202; --pw: #FFFFFF; --pp: #5939A8;
    --pgreen: #099B5A; --pred: #BA0E13;
    --gd: #1A1A1A; --gm: #3A3A3A;
}
html,body,.stApp { background:#020202!important; font-family:'Poppins',sans-serif!important; color:#e2e8f0; }
::-webkit-scrollbar{width:5px} ::-webkit-scrollbar-thumb{background:var(--pg);border-radius:99px}
.hero{text-align:center;padding:2.5rem 1rem 1.8rem;background:radial-gradient(ellipse 80% 60% at 50% 0%,rgba(255,204,63,.12) 0%,transparent 70%);border-bottom:2px solid var(--pg);margin-bottom:1.5rem}
.hero-eye{font-size:.7rem;font-weight:600;letter-spacing:.22em;color:var(--pg);text-transform:uppercase;margin-bottom:.5rem}
.hero-title{font-family:'Playfair Display',serif;font-size:2.5rem;font-weight:700;color:#fff;margin-bottom:.3rem}
.hero-title span{color:var(--pg)} .hero-sub{color:#64748b;font-size:.88rem;max-width:600px;margin:0 auto}
.glass{background:#1A1A1A;border:1px solid #2A2A2A;border-radius:12px;padding:16px;margin-bottom:14px;box-shadow:0 0 0 1px rgba(255,204,63,.05);transition:transform .22s,border-color .22s;display:flex;flex-direction:column;height:100%}
.glass:hover{transform:translateY(-3px);border-color:rgba(255,204,63,.28)}
.cpb{display:inline-flex;align-items:center;gap:4px;padding:3px 9px;border-radius:999px;font-size:.68rem;font-weight:600;margin-bottom:8px;width:fit-content;text-transform:uppercase;letter-spacing:.06em}
.ca{font-size:.95rem;font-weight:700;color:#f1f5f9;margin-bottom:2px}
.cs{font-size:.78rem;color:#94a3b8;margin-bottom:10px;line-height:1.6;display:-webkit-box;-webkit-line-clamp:4;-webkit-box-orient:vertical;overflow:hidden;border-left:3px solid var(--pg);padding-left:9px;flex-grow:1}
.cd{display:inline-flex;align-items:center;gap:4px;background:rgba(255,204,63,.08);color:var(--pg);padding:3px 9px;border-radius:999px;font-weight:600;font-size:.7rem;border:1px solid rgba(255,204,63,.2);margin-bottom:9px;width:fit-content}
.bs{display:inline-flex;align-items:center;gap:3px;padding:3px 9px;border-radius:999px;font-size:.75rem;font-weight:600;margin-right:5px;margin-bottom:3px}
.sp{background:rgba(9,155,90,.12);color:#22c55e;border:1px solid rgba(9,155,90,.2)}
.sn{background:rgba(186,14,19,.12);color:#ef4444;border:1px solid rgba(186,14,19,.2)}
.sne{background:rgba(100,116,139,.12);color:#94a3b8;border:1px solid rgba(100,116,139,.2)}
.sc{background:rgba(255,154,63,.12);color:var(--po);border:1px solid rgba(255,154,63,.25)}
.sb{background:rgba(255,204,63,.1);color:var(--pg);border:1px solid rgba(255,204,63,.25)}
.clb{display:inline-flex;align-items:center;justify-content:center;width:100%;padding:7px 0;border-radius:8px;text-decoration:none;font-weight:600;font-size:.8rem;transition:filter .2s,transform .15s;margin-top:4px;color:#fff!important}
.clb:hover{filter:brightness(1.15);transform:scale(1.01)}
.kpi-row{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:22px}
.kpi-card{background:#1A1A1A;border:1px solid #2A2A2A;border-radius:12px;padding:16px 18px;box-shadow:0 0 0 1px rgba(255,204,63,.05)}
.kpi-lbl{font-size:.67rem;color:#64748b;text-transform:uppercase;letter-spacing:.1em;margin-bottom:5px}
.kpi-val{font-size:1.9rem;font-weight:700;color:var(--pg);line-height:1.1}
.kpi-sub{font-size:.7rem;color:#475569;margin-top:3px}
.bps-wrap{background:#1A1A1A;border:1px solid #2A2A2A;border-radius:12px;padding:26px 20px;text-align:center;box-shadow:0 0 0 1px rgba(255,204,63,.06)}
.bps-score{font-size:3.5rem;font-weight:700;color:var(--pg);line-height:1}
.bps-label{font-size:.7rem;color:#64748b;text-transform:uppercase;letter-spacing:.12em;margin-top:5px}
.sh{font-size:1.05rem;font-weight:700;color:#f1f5f9;margin-bottom:.9rem;display:flex;align-items:center;gap:7px;padding-bottom:7px;border-bottom:1px solid #2A2A2A}
.sa{display:inline-block;width:36px;height:3px;background:var(--pg);border-radius:2px;margin-right:4px}
.stTabs [data-baseweb="tab-list"]{background:#1A1A1A!important;border-bottom:1px solid #2A2A2A!important}
.stTabs [data-baseweb="tab"]{background:transparent!important;font-size:.88rem;font-weight:600;padding:9px 14px;color:#64748b!important}
.stTabs [data-baseweb="tab"][aria-selected="true"]{color:var(--pg)!important;border-bottom:2px solid var(--pg)!important;background:rgba(255,204,63,.04)!important}
div.stButton>button{background:linear-gradient(135deg,var(--pg),var(--po))!important;color:#020202!important;font-weight:700!important;font-size:.93rem!important;border:none!important;border-radius:9px!important;padding:.72rem 1.1rem!important;box-shadow:0 4px 16px rgba(255,204,63,.25);font-family:'Poppins',sans-serif!important;transition:all .2s}
div.stButton>button:hover{transform:translateY(-2px)!important;box-shadow:0 8px 24px rgba(255,204,63,.38)!important}
div[data-baseweb="input"]>div,div[data-baseweb="select"]>div{background:rgba(26,26,26,.9)!important;border:1px solid #3A3A3A!important;border-radius:8px!important}
section[data-testid="stSidebar"]{background:#0A0A0A!important;border-right:1px solid #2A2A2A}
div.stDownloadButton>button{background:rgba(26,26,26,.9)!important;border:1px solid rgba(255,204,63,.3)!important;color:var(--pg)!important;border-radius:8px!important;font-weight:600!important}
div.stDownloadButton>button:hover{border-color:var(--pg)!important;background:rgba(255,204,63,.08)!important}
.pb-wrap{margin-bottom:12px;padding:11px;background:#1A1A1A;border:1px solid #2A2A2A;border-radius:10px}
.pb-lbl{font-size:.7rem;color:#94a3b8;margin-bottom:4px;display:flex;justify-content:space-between}
.pb-track{background:#2A2A2A;border-radius:4px;height:6px;overflow:hidden}
.pb-fill{height:100%;border-radius:4px}
@media(max-width:900px){.kpi-row{grid-template-columns:repeat(2,1fr)}.hero-title{font-size:1.7rem}}
</style>
""", unsafe_allow_html=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# UTILITY FUNCTIONS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _parse_ts(item):
    for key in ("posted_at","createdAt","created_at","postedAtISO","publishedAt",
                "date","timestamp","created","publishDate","datePublished"):
        val = item.get(key)
        if val is None: continue
        if isinstance(val, dict):
            ts_ms = val.get("timestamp")
            if ts_ms:
                try:
                    ts = int(ts_ms)
                    if ts > 1_000_000_000_000: ts //= 1000
                    return datetime.fromtimestamp(ts, tz=timezone.utc)
                except: pass
            ds = val.get("date","")
            if ds:
                for fmt in ("%Y-%m-%d %H:%M:%S","%Y-%m-%dT%H:%M:%S","%Y-%m-%d"):
                    try: return datetime.strptime(ds, fmt).replace(tzinfo=timezone.utc)
                    except: pass
            continue
        raw = str(val).strip()
        if not raw: continue
        if raw.isdigit():
            ts = int(raw)
            if ts > 1_000_000_000_000: ts //= 1000
            try: return datetime.fromtimestamp(ts, tz=timezone.utc)
            except: pass
            continue
        for fmt in ("%Y-%m-%dT%H:%M:%S.%fZ","%Y-%m-%dT%H:%M:%SZ",
                    "%a %b %d %H:%M:%S %z %Y","%Y-%m-%dT%H:%M:%S",
                    "%Y-%m-%d %H:%M:%S","%Y-%m-%d"):
            try:
                dt = datetime.strptime(raw, fmt)
                return dt.replace(tzinfo=timezone.utc) if not dt.tzinfo else dt
            except: pass
    return None


def _filter_time(posts, period, custom_dates):
    now = datetime.now(timezone.utc)
    cutoffs = {"today": datetime.now(IST).replace(hour=0,minute=0,second=0,microsecond=0),
               "past-24h": now - timedelta(hours=24),
               "past-week": now - timedelta(days=7),
               "past-month": now - timedelta(days=30)}
    if period in cutoffs:
        cutoff = cutoffs[period]
        return [p for p in posts if not p.get("PostedDT") or p["PostedDT"] >= cutoff]
    if period == "custom" and custom_dates:
        cd = list(custom_dates)
        sd, ed = cd[0], cd[-1]
        start = datetime(sd.year,sd.month,sd.day,0,0,0,tzinfo=IST)
        end   = datetime(ed.year,ed.month,ed.day,23,59,59,tzinfo=IST)
        return [p for p in posts if p.get("PostedDT") and start <= p["PostedDT"] <= end]
    return posts


def _sentiment(text):
    t = text.lower()
    pos_w = ["great","excellent","amazing","good","best","love","congrats","impressive","top","recommended",
              "proud","brilliant","fantastic","superb","helpful","outstanding","selected","placed","placement",
              "got into","joined","accepted","offer","gsoc","lxf","c4gt","won","award","rank"]
    neg_w = ["bad","worst","scam","fake","fraud","waste","terrible","awful","poor","disappointed",
              "useless","misleading","avoid","overrated","regret","failed","rejected","shut","closed","lied"]
    p = sum(1 for w in pos_w if w in t)
    n = sum(1 for w in neg_w if w in t)
    return "Positive" if p > n else "Negative" if n > p else "Neutral"


def _brand(text):
    return "Yes" if any(k in text.lower() for k in BRAND_KEYWORDS) else "No"


def _competitor(text):
    found = [c.title() for c in COMPETITORS if c in text.lower()]
    return ", ".join(found) if found else "None"


def _fmt_dt(dt):
    if not dt: return "", ""
    l = dt.astimezone(IST)
    return l.strftime("%d %b %Y"), l.strftime("%I:%M %p IST")


def _sent_badge(s):
    cls = {"Positive":"sp","Negative":"sn","Neutral":"sne"}.get(s,"sne")
    ic  = {"Positive":"▲","Negative":"▼","Neutral":"●"}.get(s,"●")
    return f'<span class="bs {cls}">{ic} {s}</span>'


def _brand_badge(v):
    return '<span class="bs sb">🎯 PST Tagged</span>' if v == "Yes" else ""


def _comp_badge(v):
    return f'<span class="bs sc">⚡ vs {v}</span>' if v and v != "None" else ""


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# APIFY RUNNER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _apify(actor_id, payload, api_token, label=""):
    try:
        r = requests.post(f"{APIFY_BASE}/acts/{actor_id}/runs",
                          json=payload, headers={"Content-Type":"application/json"},
                          params={"token": api_token}, timeout=30)
        r.raise_for_status()
        run_id = r.json()["data"]["id"]
    except Exception as e:
        st.error(f"❌ Apify start error: {e}"); return []

    prog = st.progress(0, text=f"⏳ {label} running…")
    for i in range(120):
        time.sleep(3)
        try:
            s = requests.get(f"{APIFY_BASE}/actor-runs/{run_id}",
                             params={"token":api_token}, timeout=15).json()["data"]["status"]
            prog.progress(min(i*2, 95), text=f"⏳ {label}: {s} ({i*3}s)…")
            if s in ("SUCCEEDED","FAILED","ABORTED","TIMED-OUT"):
                break
        except: pass

    prog.progress(100, text="✅ Done")
    if s != "SUCCEEDED":
        st.error(f"Run ended: {s}"); return []
    try:
        d = requests.get(f"{APIFY_BASE}/actor-runs/{run_id}/dataset/items",
                         params={"token":api_token}, timeout=30)
        return d.json()
    except Exception as e:
        st.error(f"❌ Results error: {e}"); return []


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SERPER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _serper(query, key, stype="search", num=20):
    ep = "https://google.serper.dev/news" if stype == "news" else "https://google.serper.dev/search"
    try:
        r = requests.post(ep, json={"q":query,"num":num,"gl":"in","hl":"en"},
                          headers={"X-API-KEY":key,"Content-Type":"application/json"}, timeout=20)
        r.raise_for_status(); return r.json()
    except Exception as e:
        st.error(f"❌ Serper error: {e}"); return {}


def _ingest_web(raw, keyword):
    out = []
    for item in (raw.get("organic",[]) + raw.get("news",[])):
        title   = item.get("title","")
        snippet = item.get("snippet","")
        text    = f"{title} {snippet}"
        out.append({
            "Date": item.get("date",""), "Title": title,
            "Snippet": snippet[:280], "Source": item.get("source", item.get("domain","Web")),
            "URL": item.get("link",""), "Position": item.get("position",99),
            "Type": "NEWS" if "publishedDate" in item else "WEB",
            "Tags Polaris": _brand(text), "Competitor": _competitor(text),
            "Sentiment": _sentiment(text), "Platform": "Web/News",
            "Scraped At": datetime.now(IST).strftime("%d %b %Y %H:%M IST"),
            "PostedDT": None,
        })
    return out


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# INGEST FUNCTIONS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _ing_linkedin(raw, kw):
    posts = []
    for item in raw:
        if not isinstance(item, dict): continue
        af = item.get("author",{}) or {}
        if not isinstance(af, dict): af = {}
        author = (af.get("name") or f"{af.get('firstName','')} {af.get('lastName','')}").strip() or item.get("authorName","Unknown")
        headline    = af.get("headline","") or item.get("authorHeadline","")
        profile_url = af.get("url","") or item.get("authorProfileUrl","")
        stats = item.get("stats",{}) or {}
        likes    = int(stats.get("total_reactions", item.get("likes", item.get("numLikes",0))) or 0)
        comments = int(stats.get("numComments", item.get("numComments",0)) or 0)
        reposts  = int(stats.get("numShares",    item.get("numShares",0)) or 0)
        activity_id = str(item.get("activity_id",""))
        post_url = str(item.get("post_url","") or item.get("url","") or item.get("postUrl",""))
        if not post_url and activity_id.isdigit():
            post_url = f"https://www.linkedin.com/feed/update/urn:li:activity:{activity_id}/"
        if not post_url: continue
        dt = _parse_ts(item); d, t = _fmt_dt(dt)
        text = str(item.get("text",""))
        posts.append({"Date":d,"Time":t,"Account":author,"Headline":headline,
            "Profile URL":profile_url,"Post URL":post_url,"Text Preview":text[:280],
            "Reactions":likes,"Comments":comments,"Reposts":reposts,
            "Est. Impressions":likes*80,"Sentiment":_sentiment(text),
            "Tags Polaris":_brand(text),"Competitor":_competitor(text),
            "Platform":"LinkedIn","Scraped At":datetime.now(IST).strftime("%d %b %Y %H:%M IST"),
            "PostedDT":dt})
    posts = list({p["Post URL"]:p for p in posts}.values())
    posts.sort(key=lambda p:p["PostedDT"] or datetime.min.replace(tzinfo=timezone.utc), reverse=True)
    return posts


def _ing_twitter(raw, kw):
    posts = []
    for item in raw:
        if not isinstance(item, dict): continue
        ai = item.get("author", item.get("user",{})) or {}
        if not isinstance(ai, dict): ai = {}
        author = ai.get("name","") or item.get("authorName","Unknown")
        handle = ai.get("username","") or ai.get("screen_name","") or ""
        if handle and not handle.startswith("@"): handle = f"@{handle}"
        profile_url = f"https://x.com/{handle.lstrip('@')}" if handle else ""
        likes    = int(item.get("likes") or item.get("likeCount") or item.get("favorite_count") or 0)
        replies  = int(item.get("replies") or item.get("reply_count") or 0)
        retweets = int(item.get("retweets") or item.get("retweet_count") or 0)
        url  = item.get("url","") or item.get("tweetUrl","")
        text = str(item.get("text","") or item.get("full_text",""))
        dt = _parse_ts(item); d, t = _fmt_dt(dt)
        posts.append({"Date":d,"Time":t,"Account":f"{author} {handle}".strip(),
            "Profile URL":profile_url,"Post URL":url,"Text Preview":text[:280],
            "Likes":likes,"Replies":replies,"Retweets":retweets,
            "Est. Impressions":likes*35,"Sentiment":_sentiment(text),
            "Tags Polaris":_brand(text),"Competitor":_competitor(text),
            "Platform":"Twitter","Scraped At":datetime.now(IST).strftime("%d %b %Y %H:%M IST"),
            "PostedDT":dt})
    return list({p["Post URL"]:p for p in posts if p["Post URL"]}.values())


def _ing_instagram(raw, kw):
    posts = []
    for item in raw:
        if not isinstance(item, dict): continue
        username = item.get("ownerUsername","") or item.get("username","") or "Unknown"
        caption  = str(item.get("caption","") or item.get("text",""))
        url      = item.get("url","") or item.get("shortCode","")
        if url and not url.startswith("http"): url = f"https://instagram.com/p/{url}"
        likes    = int(item.get("likesCount","0") or item.get("likes",0) or 0)
        comments = int(item.get("commentsCount","0") or item.get("comments",0) or 0)
        dt = _parse_ts(item); d, t = _fmt_dt(dt)
        posts.append({"Date":d,"Time":t,"Account":f"@{username}",
            "Profile URL":f"https://instagram.com/{username}","Post URL":url,
            "Caption":caption[:280],"Likes":likes,"Comments":comments,
            "Est. Impressions":likes*10,"Sentiment":_sentiment(caption),
            "Tags Polaris":_brand(caption),"Competitor":_competitor(caption),
            "Platform":"Instagram","Scraped At":datetime.now(IST).strftime("%d %b %Y %H:%M IST"),
            "PostedDT":dt})
    return posts


def _ing_facebook(raw, kw):
    posts = []
    for item in raw:
        if not isinstance(item, dict): continue
        user   = item.get("user",{}) or {}
        author = (user.get("name","") if isinstance(user,dict) else str(user)) or item.get("pageName","Unknown")
        text   = str(item.get("text","") or item.get("message",""))
        url    = item.get("url","") or item.get("link","")
        likes  = int(item.get("likes","0") or 0)
        shares = int(item.get("shares","0") or 0)
        comms  = int(item.get("comments","0") or 0)
        dt = _parse_ts(item); d, t = _fmt_dt(dt)
        posts.append({"Date":d,"Time":t,"Account":author,"Post URL":url,
            "Text Preview":text[:280],"Likes":likes,"Shares":shares,"Comments":comms,
            "Est. Impressions":likes*20,"Sentiment":_sentiment(text),
            "Tags Polaris":_brand(text),"Competitor":_competitor(text),
            "Platform":"Facebook","Scraped At":datetime.now(IST).strftime("%d %b %Y %H:%M IST"),
            "PostedDT":dt})
    return posts


def _ing_quora(raw, kw):
    posts = []
    for item in raw:
        if not isinstance(item, dict): continue
        question = item.get("question","") or item.get("title","")
        answer   = str(item.get("answer","") or item.get("content","") or item.get("text",""))
        url      = item.get("url","") or item.get("questionUrl","")
        author   = item.get("author","") or item.get("authorName","Anonymous")
        if isinstance(author, dict): author = author.get("name","Anonymous")
        upvotes  = int(item.get("upvotes","0") or item.get("likes",0) or 0)
        views    = int(item.get("views","0") or 0)
        dt = _parse_ts(item); d, t = _fmt_dt(dt)
        text_all = f"{question} {answer}"
        posts.append({"Date":d,"Time":t,"Author":author,"Question":question[:200],
            "Answer Preview":answer[:280],"URL":url,"Upvotes":upvotes,"Views":views,
            "Sentiment":_sentiment(text_all),"Tags Polaris":_brand(text_all),
            "Competitor":_competitor(text_all),"Platform":"Quora",
            "Scraped At":datetime.now(IST).strftime("%d %b %Y %H:%M IST"),"PostedDT":dt})
    return posts


def _ing_reddit(raw, kw):
    posts = []
    for item in raw:
        if not isinstance(item, dict): continue
        title     = item.get("title","") or item.get("postTitle","")
        body      = str(item.get("body","") or item.get("selftext","") or item.get("text",""))
        url       = item.get("url","") or item.get("postUrl","")
        author    = item.get("author","") or item.get("authorName","u/unknown")
        subreddit = item.get("subreddit","") or item.get("communityName","")
        upvotes   = int(item.get("upVotes","0") or item.get("score",0) or 0)
        comms     = int(item.get("numberOfComments","0") or item.get("num_comments",0) or 0)
        dt = _parse_ts(item); d, t = _fmt_dt(dt)
        text_all = f"{title} {body}"
        posts.append({"Date":d,"Time":t,"Author":f"u/{author}".replace("u/u/","u/"),
            "Subreddit":f"r/{subreddit}".replace("r/r/","r/"),
            "Title":title[:200],"Body Preview":body[:280],"URL":url,
            "Upvotes":upvotes,"Comments":comms,"Est. Impressions":upvotes*40,
            "Sentiment":_sentiment(text_all),"Tags Polaris":_brand(text_all),
            "Competitor":_competitor(text_all),"Platform":"Reddit",
            "Scraped At":datetime.now(IST).strftime("%d %b %Y %H:%M IST"),"PostedDT":dt})
    return posts


def _ing_youtube(raw, kw):
    videos, comments = [], []
    for item in raw:
        if not isinstance(item, dict): continue
        if "authorText" in item and "videoId" in item:
            text = str(item.get("text","") or item.get("commentText",""))
            dt = _parse_ts(item); d, _ = _fmt_dt(dt)
            comments.append({"Date":d,"Author":item.get("authorText","Anon"),
                "Comment":text[:300],"Video URL":f"https://youtube.com/watch?v={item.get('videoId','')}",
                "Likes":int(item.get("voteStatus","0") or item.get("likesCount",0) or 0),
                "Sentiment":_sentiment(text),"Tags Polaris":_brand(text),
                "Competitor":_competitor(text),"Platform":"YouTube",
                "Scraped At":datetime.now(IST).strftime("%d %b %Y %H:%M IST"),"PostedDT":dt})
            continue
        title  = item.get("title","") or item.get("videoTitle","")
        desc   = str(item.get("description","") or item.get("videoDescription",""))
        chan   = item.get("channelName","") or item.get("channel","") or item.get("uploaderName","")
        vid_id = item.get("id","") or item.get("videoId","")
        vid_url= item.get("url","") or (f"https://youtube.com/watch?v={vid_id}" if vid_id else "")
        views  = int(item.get("viewCount","0") or item.get("views",0) or 0)
        likes  = int(item.get("likes","0") or item.get("likeCount",0) or 0)
        comms  = int(item.get("commentsCount","0") or item.get("commentCount",0) or 0)
        dt = _parse_ts(item); d, _ = _fmt_dt(dt)
        text_all = f"{title} {desc}"
        videos.append({"Date":d,"Channel":chan,"Channel URL":item.get("channelUrl",""),
            "Title":title[:200],"Description Preview":desc[:300],"Video URL":vid_url,
            "Views":views,"Likes":likes,"Comments":comms,
            "Sentiment":_sentiment(text_all),"Tags Polaris":_brand(text_all),
            "Competitor":_competitor(text_all),"Platform":"YouTube",
            "Scraped At":datetime.now(IST).strftime("%d %b %Y %H:%M IST"),"PostedDT":dt})
    return videos, comments


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# BPS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _bps(ss):
    weights = {"linkedin":20,"twitter":15,"instagram":10,"facebook":8,
               "quora":15,"reddit":12,"youtube":12,"web":8}
    keys = {"linkedin":"posts_linkedin","twitter":"posts_twitter","instagram":"posts_instagram",
            "facebook":"posts_facebook","quora":"posts_quora","reddit":"posts_reddit",
            "youtube":"posts_youtube_videos","web":"posts_web"}
    pscores = {}
    for k, sk in keys.items():
        data = ss.get(sk,[])
        if not data: pscores[k] = 0; continue
        brand = sum(1 for d in data if d.get("Tags Polaris")=="Yes")
        pos   = sum(1 for d in data if d.get("Sentiment")=="Positive")
        pscores[k] = min(min(len(data)*3,40) + min(brand*5,30) + min(pos*4,30), 100)
    weighted = sum(pscores[k]*weights[k] for k in weights) / sum(weights.values())
    return round(weighted), pscores


def _grade(score):
    if score >= 80: return "🟢 Excellent", "#22c55e"
    if score >= 60: return "🟡 Good", "#FFCC3F"
    if score >= 40: return "🟠 Building", "#FF9A3F"
    return "🔴 Needs Work", "#BA0E13"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CARD RENDERERS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _cards(posts, platform, url_key="Post URL", text_key="Text Preview",
           author_key="Account", stat_keys=("Reactions",)):
    if not posts:
        st.info("No results yet — hit the Scrape button above."); return
    color = PLATFORM_COLORS.get(platform,"#888")
    for i in range(0, min(len(posts), 60), 3):
        cols = st.columns(3)
        for j, col in enumerate(cols):
            if i+j >= len(posts): break
            p = posts[i+j]
            url  = p.get(url_key,"#") or "#"
            text = p.get(text_key,"")
            auth = p.get(author_key,"")
            stats_html = ""
            for sk in stat_keys:
                v = p.get(sk,0)
                if v: stats_html += f'<span class="bs" style="background:rgba(255,204,63,.08);color:#FFCC3F;border:1px solid rgba(255,204,63,.2)">📊 {sk}: {v:,}</span>'
            imp = p.get("Est. Impressions",0)
            if imp: stats_html += f'<span class="bs" style="background:rgba(9,155,90,.08);color:#22c55e;border:1px solid rgba(9,155,90,.2)">👁 ~{imp:,}</span>'
            dt_str = f"{p.get('Date','')} {p.get('Time','')}".strip() or "—"
            col.markdown(f"""
<div class="glass">
  <span class="cpb" style="background:rgba(0,0,0,.35);color:{color};border:1px solid {color}44">● {platform}</span>
  <div class="ca">{html.escape(str(auth))}</div>
  <div class="cd">📅 {dt_str}</div>
  <div class="cs">{html.escape(str(text))}</div>
  <div style="margin-bottom:9px">{stats_html}</div>
  <div style="margin-bottom:9px">{_sent_badge(p.get("Sentiment","Neutral"))}{_brand_badge(p.get("Tags Polaris","No"))}{_comp_badge(p.get("Competitor","None"))}</div>
  <a href="{url}" target="_blank" class="clb" style="background:{color}cc">View Post ↗</a>
</div>""", unsafe_allow_html=True)


def _quora_cards(posts):
    if not posts: st.info("No results yet."); return
    color = PLATFORM_COLORS["Quora"]
    for i in range(0, min(len(posts), 60), 3):
        cols = st.columns(3)
        for j, col in enumerate(cols):
            if i+j >= len(posts): break
            p = posts[i+j]
            col.markdown(f"""
<div class="glass">
  <span class="cpb" style="background:rgba(0,0,0,.35);color:{color};border:1px solid {color}44">● Quora</span>
  <div class="ca">{html.escape(str(p.get("Author","")))} </div>
  <div class="cd">📅 {p.get("Date","—")}</div>
  <div class="cs"><b>Q:</b> {html.escape(str(p.get("Question",""))[:200])}</div>
  <div class="cs">{html.escape(str(p.get("Answer Preview",""))[:250])}</div>
  <div style="margin-bottom:9px">
    <span class="bs" style="background:rgba(255,204,63,.08);color:#FFCC3F;border:1px solid rgba(255,204,63,.2)">▲ {p.get("Upvotes",0):,} Upvotes</span>
    {_sent_badge(p.get("Sentiment","Neutral"))}{_brand_badge(p.get("Tags Polaris","No"))}{_comp_badge(p.get("Competitor","None"))}
  </div>
  <a href="{p.get('URL','#')}" target="_blank" class="clb" style="background:{color}cc">View on Quora ↗</a>
</div>""", unsafe_allow_html=True)


def _reddit_cards(posts):
    if not posts: st.info("No results yet."); return
    color = PLATFORM_COLORS["Reddit"]
    for i in range(0, min(len(posts), 60), 3):
        cols = st.columns(3)
        for j, col in enumerate(cols):
            if i+j >= len(posts): break
            p = posts[i+j]
            col.markdown(f"""
<div class="glass">
  <span class="cpb" style="background:rgba(0,0,0,.35);color:{color};border:1px solid {color}44">● Reddit · {html.escape(str(p.get("Subreddit","")))} </span>
  <div class="ca">{html.escape(str(p.get("Author","")))} </div>
  <div class="cd">📅 {p.get("Date","—")}</div>
  <div class="cs"><b>{html.escape(str(p.get("Title",""))[:120])}</b></div>
  <div class="cs">{html.escape(str(p.get("Body Preview",""))[:240])}</div>
  <div style="margin-bottom:9px">
    <span class="bs" style="background:rgba(255,204,63,.08);color:#FFCC3F;border:1px solid rgba(255,204,63,.2)">▲ {p.get("Upvotes",0):,}</span>
    <span class="bs" style="background:rgba(100,116,139,.08);color:#94a3b8;border:1px solid rgba(100,116,139,.2)">💬 {p.get("Comments",0):,}</span>
    {_sent_badge(p.get("Sentiment","Neutral"))}{_brand_badge(p.get("Tags Polaris","No"))}
  </div>
  <a href="{p.get('URL','#')}" target="_blank" class="clb" style="background:{color}cc">View Thread ↗</a>
</div>""", unsafe_allow_html=True)


def _yt_cards(videos, comments):
    t1, t2 = st.tabs(["▶️ Videos + Descriptions", "💬 Comments"])
    with t1:
        if not videos: st.info("No videos yet."); 
        else:
            color = "#FF0000"
            for i in range(0, min(len(videos),30), 3):
                cols = st.columns(3)
                for j, col in enumerate(cols):
                    if i+j >= len(videos): break
                    v = videos[i+j]
                    col.markdown(f"""
<div class="glass">
  <span class="cpb" style="background:rgba(0,0,0,.35);color:{color};border:1px solid {color}44">● YouTube</span>
  <div class="ca">{html.escape(str(v.get("Channel","")))} </div>
  <div class="cd">📅 {v.get("Date","—")}</div>
  <div class="cs"><b>{html.escape(str(v.get("Title",""))[:150])}</b></div>
  <div class="cs">{html.escape(str(v.get("Description Preview",""))[:250])}</div>
  <div style="margin-bottom:9px">
    <span class="bs" style="background:rgba(255,204,63,.08);color:#FFCC3F;border:1px solid rgba(255,204,63,.2)">👁 {v.get("Views",0):,}</span>
    <span class="bs" style="background:rgba(251,113,133,.08);color:#fb7185;border:1px solid rgba(251,113,133,.15)">♥ {v.get("Likes",0):,}</span>
    {_sent_badge(v.get("Sentiment","Neutral"))}{_brand_badge(v.get("Tags Polaris","No"))}{_comp_badge(v.get("Competitor","None"))}
  </div>
  <a href="{v.get('Video URL','#')}" target="_blank" class="clb" style="background:{color}cc">Watch Video ↗</a>
</div>""", unsafe_allow_html=True)
    with t2:
        if not comments: st.info("No comments yet — enable 'Fetch comments' toggle before scraping.")
        else:
            color = "#FF0000"
            for i in range(0, min(len(comments),60), 3):
                cols = st.columns(3)
                for j, col in enumerate(cols):
                    if i+j >= len(comments): break
                    c = comments[i+j]
                    col.markdown(f"""
<div class="glass">
  <span class="cpb" style="background:rgba(0,0,0,.35);color:{color};border:1px solid {color}44">● YT Comment</span>
  <div class="ca">{html.escape(str(c.get("Author","")))} </div>
  <div class="cd">📅 {c.get("Date","—")}</div>
  <div class="cs">{html.escape(str(c.get("Comment",""))[:280])}</div>
  <div style="margin-bottom:9px">
    <span class="bs" style="background:rgba(251,113,133,.08);color:#fb7185;border:1px solid rgba(251,113,133,.15)">♥ {c.get("Likes",0):,}</span>
    {_sent_badge(c.get("Sentiment","Neutral"))}{_brand_badge(c.get("Tags Polaris","No"))}
  </div>
  <a href="{c.get('Video URL','#')}" target="_blank" class="clb" style="background:{color}cc">View Video ↗</a>
</div>""", unsafe_allow_html=True)


def _web_cards(posts):
    if not posts: st.info("No results yet — hit Search Web & News above."); return
    color = PLATFORM_COLORS["Web/News"]
    for i in range(0, min(len(posts),60), 3):
        cols = st.columns(3)
        for j, col in enumerate(cols):
            if i+j >= len(posts): break
            p = posts[i+j]
            col.markdown(f"""
<div class="glass">
  <span class="cpb" style="background:rgba(0,0,0,.35);color:{color};border:1px solid {color}44">
    ● {p.get("Type","WEB")} · {html.escape(str(p.get("Source",""))[:28])}
  </span>
  <div class="ca">{html.escape(str(p.get("Title",""))[:100])}</div>
  <div class="cd">📅 {p.get("Date","—")} · Rank #{p.get("Position","?")}</div>
  <div class="cs">{html.escape(str(p.get("Snippet",""))[:240])}</div>
  <div style="margin-bottom:9px">{_sent_badge(p.get("Sentiment","Neutral"))}{_brand_badge(p.get("Tags Polaris","No"))}{_comp_badge(p.get("Competitor","None"))}</div>
  <a href="{p.get('URL','#')}" target="_blank" class="clb" style="background:{color}cc">Read Article ↗</a>
</div>""", unsafe_allow_html=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SCRAPER CONTROLS HELPER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _controls(pk, default_kw="Polaris School of Technology"):
    col1, col2 = st.columns([2.5, 1])
    with col1: kw = st.text_input("Keyword / Brand / Hashtag", value=default_kw, key=f"kw_{pk}")
    with col2: period = st.selectbox("Time Window", ["past-week","today","past-24h","past-month","custom"], key=f"pd_{pk}")
    custom_dates = None
    if period == "custom":
        custom_dates = st.date_input("Date Range",
            value=(datetime.today().date()-timedelta(days=7), datetime.today().date()),
            max_value=datetime.today().date(), key=f"dt_{pk}")
        if not isinstance(custom_dates,(list,tuple)): custom_dates = [custom_dates]
    c1, c2 = st.columns([2,1])
    with c1: max_items = st.slider("Max Items", 10, 200, 50, step=10, key=f"mx_{pk}")
    with c2: st.metric("Est. Cost", f"${max_items * COSTS.get(pk, 0.005):.3f}")
    return kw, period, custom_dates, max_items


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# EXCEL EXPORT
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _excel(ss):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    gold_fill = PatternFill("solid", fgColor="FFCC3F")
    hdr_font  = Font(bold=True, color="020202", name="Calibri", size=10)

    def _sheet(name, rows, cols):
        if not rows: return
        ws = wb.create_sheet(title=name[:31])
        ws.append(cols)
        for cell in ws[1]:
            cell.fill = gold_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in rows:
            ws.append([str(row.get(c,"")) for c in cols])
        for cc in ws.columns:
            ml = max(len(str(c.value or "")) for c in cc)
            ws.column_dimensions[get_column_letter(cc[0].column)].width = min(ml+4, 52)
        ws.freeze_panes = "A2"

    sheets = [
        ("LinkedIn",   "posts_linkedin",  ["Date","Time","Account","Headline","Post URL","Text Preview","Reactions","Comments","Reposts","Est. Impressions","Sentiment","Tags Polaris","Competitor","Scraped At"]),
        ("Twitter",    "posts_twitter",   ["Date","Time","Account","Post URL","Text Preview","Likes","Replies","Retweets","Est. Impressions","Sentiment","Tags Polaris","Competitor","Scraped At"]),
        ("Instagram",  "posts_instagram", ["Date","Time","Account","Post URL","Caption","Likes","Comments","Est. Impressions","Sentiment","Tags Polaris","Competitor","Scraped At"]),
        ("Facebook",   "posts_facebook",  ["Date","Time","Account","Post URL","Text Preview","Likes","Shares","Comments","Est. Impressions","Sentiment","Tags Polaris","Competitor","Scraped At"]),
        ("Quora",      "posts_quora",     ["Date","Time","Author","Question","Answer Preview","URL","Upvotes","Views","Sentiment","Tags Polaris","Competitor","Scraped At"]),
        ("Reddit",     "posts_reddit",    ["Date","Time","Author","Subreddit","Title","Body Preview","URL","Upvotes","Comments","Est. Impressions","Sentiment","Tags Polaris","Competitor","Scraped At"]),
        ("YT-Videos",  "posts_youtube_videos", ["Date","Channel","Title","Description Preview","Video URL","Views","Likes","Comments","Sentiment","Tags Polaris","Competitor","Scraped At"]),
        ("YT-Comments","posts_youtube_comments", ["Date","Author","Comment","Video URL","Likes","Sentiment","Tags Polaris","Competitor","Scraped At"]),
        ("Web-News",   "posts_web",       ["Date","Title","Snippet","Source","URL","Position","Type","Sentiment","Tags Polaris","Competitor","Scraped At"]),
    ]
    for sname, key, cols in sheets:
        _sheet(sname, ss.get(key,[]), cols)

    ws_sum = wb.create_sheet(title="Summary", index=0)
    ws_sum.append(["Platform","Total","Tags Polaris","Positive","Negative","Neutral","Competitor Overlap"])
    for cell in ws_sum[1]: cell.fill = gold_fill; cell.font = hdr_font
    for sname, key, _ in sheets:
        data = ss.get(key,[])
        ws_sum.append([sname, len(data),
            sum(1 for d in data if d.get("Tags Polaris")=="Yes"),
            sum(1 for d in data if d.get("Sentiment")=="Positive"),
            sum(1 for d in data if d.get("Sentiment")=="Negative"),
            sum(1 for d in data if d.get("Sentiment")=="Neutral"),
            sum(1 for d in data if d.get("Competitor","None")!="None")])
    for cc in ws_sum.columns:
        ml = max(len(str(c.value or "")) for c in cc)
        ws_sum.column_dimensions[get_column_letter(cc[0].column)].width = min(ml+4, 40)
    ws_sum.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SESSION STATE INIT
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
for _k in ["posts_linkedin","posts_twitter","posts_instagram","posts_facebook",
           "posts_quora","posts_reddit","posts_youtube_videos","posts_youtube_comments","posts_web"]:
    if _k not in st.session_state: st.session_state[_k] = []


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# HERO
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
st.markdown("""
<div class="hero">
  <div class="hero-eye">Polaris School of Technology · Brand Intelligence Suite</div>
  <div class="hero-title">🎯 <span>Brand</span> Monitor</div>
  <div class="hero-sub">Track every mention — LinkedIn · Twitter · Instagram · Facebook · Quora · Reddit · YouTube · Web & News</div>
</div>
""", unsafe_allow_html=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SIDEBAR
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
with st.sidebar:
    st.markdown("## 🎯 Polaris Brand Monitor")
    st.markdown("---")
    api_token = st.text_input("🔑 Apify API Token", value=os.getenv("APIFY_TOKEN",""),
        type="password", placeholder="apify_api_xxxx…",
        help="apify.com → Settings → Integrations → API Token")
    if api_token: st.success("✅ Apify set")
    else: st.warning("⚠️ Required for social tabs")
    st.markdown("---")
    serper_key = st.text_input("🌐 Serper API Key", value=os.getenv("SERPER_API_KEY",""),
        type="password", placeholder="xxxx…",
        help="serper.dev → Dashboard → API Key — 2,500 free/month")
    if serper_key: st.success("✅ Serper set")
    else: st.caption("⚠️ Required for Web & News tab")
    st.markdown("---")
    st.markdown("**🏷️ Brand Keywords (auto)**")
    st.code("polaris · polariscampus\npolaris school of technology\npst bengaluru", language=None)
    st.markdown("**⚡ Competitors Tracked**")
    st.code("Scaler · Newton School\nupGrad · Great Learning\nIIT Bombay/Delhi/Madras", language=None)
    st.markdown("---")
    st.markdown("**💰 Apify Cost Guide**")
    st.caption("LinkedIn: ~$0.005/post\nTwitter: ~$0.00015/tweet\nInstagram: ~$0.004/post\nFacebook: ~$0.004/post\nQuora: ~$0.005/run\nReddit: ~$0.005/run\nYouTube: ~$0.01/video")
    st.markdown("---")
    st.caption("[Streamlit Deploy Guide ↗](https://share.streamlit.io)")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# TABS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
tabs = st.tabs([
    "🏠 Overview", "🔗 LinkedIn", "𝕏 Twitter",
    "📸 Instagram", "👥 Facebook", "❓ Quora",
    "🤖 Reddit", "▶️ YouTube", "🌐 Web & News",
    "📥 Export", "ℹ️ Guide",
])
tab_ov, tab_li, tab_tw, tab_ig, tab_fb, tab_qr, tab_rd, tab_yt, tab_wb, tab_ex, tab_hl = tabs


# ━━ OVERVIEW ━━
with tab_ov:
    score, pscores = _bps(st.session_state)
    grade_txt, grade_col = _grade(score)

    col_bps, col_kpis = st.columns([1, 3])
    with col_bps:
        total_all = sum(len(st.session_state.get(k,[])) for k in
            ["posts_linkedin","posts_twitter","posts_instagram","posts_facebook",
             "posts_quora","posts_reddit","posts_youtube_videos","posts_web"])
        st.markdown(f"""
<div class="bps-wrap">
  <div class="bps-score">{score}</div>
  <div class="bps-label">Brand Presence Score</div>
  <div style="font-size:1rem;font-weight:600;color:{grade_col};margin-top:7px">{grade_txt}</div>
  <div style="font-size:.7rem;color:#475569;margin-top:10px">{total_all:,} total mentions tracked</div>
</div>""", unsafe_allow_html=True)

    with col_kpis:
        all_data = []
        for k in ["posts_linkedin","posts_twitter","posts_instagram","posts_facebook",
                  "posts_quora","posts_reddit","posts_youtube_videos","posts_web"]:
            all_data.extend(st.session_state.get(k,[]))
        total     = len(all_data)
        brand_cnt = sum(1 for d in all_data if d.get("Tags Polaris")=="Yes")
        pos_cnt   = sum(1 for d in all_data if d.get("Sentiment")=="Positive")
        comp_cnt  = sum(1 for d in all_data if d.get("Competitor","None")!="None")
        st.markdown(f"""
<div class="kpi-row">
  <div class="kpi-card"><div class="kpi-lbl">Total Mentions</div><div class="kpi-val">{total:,}</div><div class="kpi-sub">across all platforms</div></div>
  <div class="kpi-card"><div class="kpi-lbl">Tags Polaris</div><div class="kpi-val">{brand_cnt:,}</div><div class="kpi-sub">{round(brand_cnt/total*100) if total else 0}% of total</div></div>
  <div class="kpi-card"><div class="kpi-lbl">Positive Sentiment</div><div class="kpi-val" style="color:#22c55e">{pos_cnt:,}</div><div class="kpi-sub">{round(pos_cnt/total*100) if total else 0}% of total</div></div>
  <div class="kpi-card"><div class="kpi-lbl">Competitor Overlap</div><div class="kpi-val" style="color:#FF9A3F">{comp_cnt:,}</div><div class="kpi-sub">comparison posts</div></div>
</div>""", unsafe_allow_html=True)

    st.markdown('<div class="sh"><span class="sa"></span>Platform Breakdown</div>', unsafe_allow_html=True)
    pm = {
        "LinkedIn":  ("posts_linkedin",  "#0A66C2"),
        "Twitter":   ("posts_twitter",   "#1DA1F2"),
        "Instagram": ("posts_instagram", "#E1306C"),
        "Facebook":  ("posts_facebook",  "#1877F2"),
        "Quora":     ("posts_quora",     "#B92B27"),
        "Reddit":    ("posts_reddit",    "#FF4500"),
        "YouTube":   ("posts_youtube_videos","#FF0000"),
        "Web/News":  ("posts_web",       "#099B5A"),
    }
    max_c = max((len(st.session_state.get(v[0],[])) for v in pm.values()), default=1) or 1
    pb_cols = st.columns(2)
    for idx, (plat, (key, color)) in enumerate(pm.items()):
        cnt = len(st.session_state.get(key,[]))
        pct = int(cnt / max_c * 100)
        bp  = sum(1 for d in st.session_state.get(key,[]) if d.get("Tags Polaris")=="Yes")
        with pb_cols[idx % 2]:
            st.markdown(f"""
<div class="pb-wrap">
  <div class="pb-lbl"><span style="color:{color};font-weight:600">● {plat}</span><span>{cnt} mentions · {bp} tagged PST</span></div>
  <div class="pb-track"><div class="pb-fill" style="width:{pct}%;background:{color}"></div></div>
</div>""", unsafe_allow_html=True)
    if total == 0:
        st.info("👆 Run scrapers in each tab first. Overview populates automatically as you collect data.")


# ━━ LINKEDIN ━━
with tab_li:
    st.markdown('<div class="sh"><span class="sa"></span>LinkedIn Mentions</div>', unsafe_allow_html=True)
    kw, period, cdt, mx = _controls("linkedin")
    st.markdown("---")
    if st.button("🚀 Scrape LinkedIn", use_container_width=True, key="btn_li"):
        if not api_token: st.error("Add Apify token in sidebar first.")
        else:
            payload = {"query": kw, "maxResults": mx,
                       "datePosted": period if period not in ("custom","today","past-24h") else "pastWeek"}
            raw = _apify(ACTORS["linkedin"], payload, api_token, "LinkedIn")
            posts = _ing_linkedin(raw, kw)
            posts = _filter_time(posts, period, cdt)
            st.session_state.posts_linkedin = posts
            st.success(f"✅ {len(posts)} LinkedIn posts")
    _cards(st.session_state.posts_linkedin, "LinkedIn", stat_keys=("Reactions","Comments","Reposts"))


# ━━ TWITTER ━━
with tab_tw:
    st.markdown('<div class="sh"><span class="sa"></span>Twitter / X Mentions</div>', unsafe_allow_html=True)
    kw, period, cdt, mx = _controls("twitter")
    st.markdown("---")
    if st.button("🚀 Scrape Twitter / X", use_container_width=True, key="btn_tw"):
        if not api_token: st.error("Add Apify token in sidebar first.")
        else:
            payload = {"searchTerms": [kw], "maxItems": mx, "queryType": "Latest"}
            raw = _apify(ACTORS["twitter"], payload, api_token, "Twitter")
            posts = _ing_twitter(raw, kw)
            posts = _filter_time(posts, period, cdt)
            st.session_state.posts_twitter = posts
            st.success(f"✅ {len(posts)} tweets")
    _cards(st.session_state.posts_twitter, "Twitter", stat_keys=("Likes","Replies","Retweets"))


# ━━ INSTAGRAM ━━
with tab_ig:
    st.markdown('<div class="sh"><span class="sa"></span>Instagram Mentions</div>', unsafe_allow_html=True)
    kw, period, cdt, mx = _controls("instagram", "#polariscampus")
    st.markdown("---")
    if st.button("🚀 Scrape Instagram", use_container_width=True, key="btn_ig"):
        if not api_token: st.error("Add Apify token in sidebar first.")
        else:
            payload = {"hashtags": [kw.lstrip("#")], "resultsLimit": mx}
            raw = _apify(ACTORS["instagram"], payload, api_token, "Instagram")
            posts = _ing_instagram(raw, kw)
            st.session_state.posts_instagram = posts
            st.success(f"✅ {len(posts)} Instagram posts")
    _cards(st.session_state.posts_instagram, "Instagram",
           url_key="Post URL", text_key="Caption", stat_keys=("Likes","Comments"))


# ━━ FACEBOOK ━━
with tab_fb:
    st.markdown('<div class="sh"><span class="sa"></span>Facebook Public Mentions</div>', unsafe_allow_html=True)
    kw, period, cdt, mx = _controls("facebook")
    st.markdown("---")
    if st.button("🚀 Scrape Facebook", use_container_width=True, key="btn_fb"):
        if not api_token: st.error("Add Apify token in sidebar first.")
        else:
            payload = {"query": kw, "maxPosts": mx}
            raw = _apify(ACTORS["facebook"], payload, api_token, "Facebook")
            posts = _ing_facebook(raw, kw)
            st.session_state.posts_facebook = posts
            st.success(f"✅ {len(posts)} Facebook posts")
    _cards(st.session_state.posts_facebook, "Facebook", stat_keys=("Likes","Shares","Comments"))


# ━━ QUORA ━━
with tab_qr:
    st.markdown('<div class="sh"><span class="sa"></span>Quora Q&A Mentions</div>', unsafe_allow_html=True)
    kw, period, cdt, mx = _controls("quora")
    st.markdown("---")
    if st.button("🚀 Scrape Quora", use_container_width=True, key="btn_qr"):
        if not api_token: st.error("Add Apify token in sidebar first.")
        else:
            payload = {"queries": [kw], "maxResults": mx}
            raw = _apify(ACTORS["quora"], payload, api_token, "Quora")
            posts = _ing_quora(raw, kw)
            st.session_state.posts_quora = posts
            st.success(f"✅ {len(posts)} Quora results")
    _quora_cards(st.session_state.posts_quora)


# ━━ REDDIT ━━
with tab_rd:
    st.markdown('<div class="sh"><span class="sa"></span>Reddit Posts & Threads</div>', unsafe_allow_html=True)
    kw, period, cdt, mx = _controls("reddit")
    st.markdown("---")
    if st.button("🚀 Scrape Reddit", use_container_width=True, key="btn_rd"):
        if not api_token: st.error("Add Apify token in sidebar first.")
        else:
            payload = {"searches":[{"query":kw,"sort":"relevance","time":"month"}], "maxItems": mx}
            raw = _apify(ACTORS["reddit"], payload, api_token, "Reddit")
            posts = _ing_reddit(raw, kw)
            st.session_state.posts_reddit = posts
            st.success(f"✅ {len(posts)} Reddit posts")
    _reddit_cards(st.session_state.posts_reddit)


# ━━ YOUTUBE ━━
with tab_yt:
    st.markdown('<div class="sh"><span class="sa"></span>YouTube — Videos, Descriptions & Comments</div>', unsafe_allow_html=True)
    kw, period, cdt, mx = _controls("youtube")
    fetch_comments = st.toggle("Also fetch YouTube comments (slower)", value=False, key="yt_com")
    st.markdown("---")
    if st.button("🚀 Scrape YouTube", use_container_width=True, key="btn_yt"):
        if not api_token: st.error("Add Apify token in sidebar first.")
        else:
            payload = {"searchKeywords":kw,"maxResults":mx,
                       "includeComments":fetch_comments,"maxComments":100 if fetch_comments else 0}
            raw = _apify(ACTORS["youtube"], payload, api_token, "YouTube")
            vids, coms = _ing_youtube(raw, kw)
            st.session_state.posts_youtube_videos = vids
            st.session_state.posts_youtube_comments = coms
            st.success(f"✅ {len(vids)} videos · {len(coms)} comments")
    _yt_cards(st.session_state.posts_youtube_videos, st.session_state.posts_youtube_comments)


# ━━ WEB & NEWS ━━
with tab_wb:
    st.markdown('<div class="sh"><span class="sa"></span>Web, News & Medium (Serper)</div>', unsafe_allow_html=True)
    c1, c2 = st.columns([2.5, 1])
    with c1: web_kw = st.text_input("Search Query", value="Polaris School of Technology", key="kw_web")
    with c2: stype = st.selectbox("Type", ["Combined","News Only","Web Only"], key="web_type")
    c3, c4 = st.columns([2, 1])
    with c3: web_num = st.slider("Results", 10, 100, 30, step=10, key="web_num")
    with c4: st.metric("Est. Cost", f"${web_num*0.001:.3f}")
    st.markdown("---")
    if st.button("🚀 Search Web & News", use_container_width=True, key="btn_wb"):
        if not serper_key: st.error("Add Serper key in sidebar first.")
        else:
            results = []
            if stype in ("Combined","Web Only"):
                results.extend(_ingest_web(_serper(web_kw, serper_key, "search", web_num), web_kw))
            if stype in ("Combined","News Only"):
                results.extend(_ingest_web(_serper(web_kw, serper_key, "news",   web_num), web_kw))
            seen, deduped = set(), []
            for r in results:
                u = r.get("URL","")
                if u not in seen: seen.add(u); deduped.append(r)
            st.session_state.posts_web = deduped
            st.success(f"✅ {len(deduped)} web/news results")
    _web_cards(st.session_state.posts_web)


# ━━ EXPORT ━━
with tab_ex:
    st.markdown('<div class="sh"><span class="sa"></span>Download Brand Intelligence Report</div>', unsafe_allow_html=True)
    counts = {
        "LinkedIn": len(st.session_state.posts_linkedin),
        "Twitter": len(st.session_state.posts_twitter),
        "Instagram": len(st.session_state.posts_instagram),
        "Facebook": len(st.session_state.posts_facebook),
        "Quora": len(st.session_state.posts_quora),
        "Reddit": len(st.session_state.posts_reddit),
        "YT Videos": len(st.session_state.posts_youtube_videos),
        "YT Comments": len(st.session_state.posts_youtube_comments),
        "Web/News": len(st.session_state.posts_web),
    }
    total_r = sum(counts.values())
    st.markdown(f"""
<div class="kpi-row" style="grid-template-columns:repeat(5,1fr)">
  {''.join(f'<div class="kpi-card"><div class="kpi-lbl">{k}</div><div class="kpi-val">{v:,}</div></div>' for k,v in counts.items() if v)}
</div>""", unsafe_allow_html=True)
    if total_r == 0:
        st.warning("No data yet — run the scrapers in each tab first, then come back here to export.")
    else:
        fname = f"Polaris_Brand_{datetime.now(IST).strftime('%d%b%Y_%H%M')}.xlsx"
        try:
            xls = _excel(st.session_state)
            st.download_button(
                f"📥 Download Excel Report — {total_r:,} records",
                data=xls, file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
            st.success(f"✅ Ready — {total_r:,} records · 10 tabs (Summary + 1 per platform)")
        except Exception as e:
            st.error(f"Export error: {e}")


# ━━ GUIDE ━━
with tab_hl:
    st.markdown("""
## 🎯 Setup & Usage Guide

---

### 🔑 API Keys You Need

| Key | Where to Get | Free Tier |
|---|---|---|
| **Apify** | apify.com → Settings → Integrations | ~$5 credit free |
| **Serper** | serper.dev → Dashboard | 2,500 searches/month |

---

### 📋 Platform Coverage

| Platform | What It Pulls | Source |
|---|---|---|
| LinkedIn | Posts, reactions, comments, reposts | Apify |
| Twitter/X | Tweets, likes, replies, retweets | Apify |
| Instagram | Posts + captions via hashtag search | Apify |
| Facebook | Public posts mentioning keyword | Apify |
| Quora | Questions + answers | Apify |
| Reddit | Posts + threads + upvotes | Apify |
| YouTube | Videos + descriptions + comments | Apify |
| Web/News | Google-indexed pages, Medium, news portals | Serper |

---

### 🧠 Automatic Intelligence on Every Card

- **Sentiment** — Positive / Neutral / Negative (keyword model)
- **Tags PST** — Flags posts explicitly naming Polaris
- **Competitor flag** — Detects Scaler, Newton, upGrad, Great Learning in same post
- **Est. Impressions** — Engagement × platform multiplier

---

### 📊 Brand Presence Score (BPS)

Weighted across all 8 platforms. Goes up with more mentions, more brand-tagged posts, and more positive sentiment.

---

### ☁️ Deploy Free on Streamlit Cloud

1. Push `app.py` + `requirements.txt` to GitHub
2. Go to [share.streamlit.io](https://share.streamlit.io) → New App
3. Under **Secrets**, add:
   ```
   APIFY_TOKEN = "apify_api_..."
   SERPER_API_KEY = "..."
   ```
4. Live URL: `https://yourname-polaris-monitor.streamlit.app`

---

### 📁 Requirements

```
streamlit>=1.28.0
pandas>=2.0.0
requests>=2.31.0
openpyxl>=3.1.0
python-dotenv>=1.0.0
```
""")
